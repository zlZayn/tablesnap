"""Workflow orchestration — ties capture, VLM analysis, and export together."""

import subprocess
import time
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from core.config import DEBOUNCE, HOTKEY, OUTPUT_DIR, VLM_OLLAMA_URL, VLM_MODEL, VLM_TIMEOUT
from core.hotkey import wait_for_hotkey
from core.output import (
    console,
    print_banner,
    print_break,
    print_err,
    print_ok,
    print_rule,
    print_stage,
    print_timing,
    print_tip,
    print_warn,
    spinner,
)
from vlm.client import OllamaClient
from export.xlsx import psv_to_xlsx
from capture.selector import capture_region


def _xlsx_shape(path: str) -> str:
    """Read xlsx shape quickly — '3r x 5c' or '?'."""
    try:
        wb = load_workbook(path)
        ws = wb.active
        if ws is not None:
            return f"{ws.max_row}r x {ws.max_column}c"
    except Exception:
        pass
    return "?"


def _analyze_and_export(
    image_bytes: bytes,
    ts: str,
) -> tuple[str | None, str, float, float]:
    """VLM analyze an image, then export the result to XLSX.

    Prints the ``vlm`` and ``export`` stages with their timings.

    Args:
        image_bytes: Raw image bytes to send to the VLM.
        ts:          Timestamp string shared by the XLSX filename.

    Returns:
        ``(xlsx_path, raw_text, vlm_seconds, export_seconds)`` where
        ``xlsx_path`` is ``None`` when the VLM returned an error or
        reported no table.
    """
    # -- VLM analysis --
    print_stage("vlm")
    t_vlm = time.perf_counter()
    client = OllamaClient(
        base_url=VLM_OLLAMA_URL,
        model=VLM_MODEL,
        timeout=VLM_TIMEOUT,
    )
    with spinner("vlm analyzing"):
        psv_text = client.analyze(image_bytes)
    t_vlm = time.perf_counter() - t_vlm
    print_timing("vlm analyze", t_vlm)

    raw_stripped = psv_text.strip()

    # -- Export --
    print_stage("export")
    t_xport = time.perf_counter()
    if raw_stripped.startswith("ERROR:") or raw_stripped == "NO_TABLE":
        xlsx_path = None
    else:
        xlsx_path = psv_to_xlsx(psv_text, timestamp=ts)
    t_xport = time.perf_counter() - t_xport
    print_timing("export", t_xport)

    return xlsx_path, raw_stripped, t_vlm, t_xport


def process_screenshot() -> None:
    """1. Capture region → 2. VLM analyze → 3. Export to XLSX."""
    try:
        # -- Step 1: Capture --
        print_stage("capture")
        t_cap = time.perf_counter()
        # Generate timestamp once — shared by screenshot and XLSX
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        try:
            image_path = capture_region(timestamp=ts)
        except Exception as exc:
            print_timing("capture", time.perf_counter() - t_cap)
            print_err(f"capture failed: {exc}")
            print_tip("Check display and permissions, then try again")
            return
        t_cap = time.perf_counter() - t_cap

        if image_path is None:
            print_timing("capture", t_cap, "cancelled")
            return

        print_timing("capture", t_cap)

        # -- Step 2+3: VLM analysis + export --
        image_bytes = Path(image_path).read_bytes()
        xlsx_path, raw_stripped, t_vlm, t_xport = _analyze_and_export(
            image_bytes, ts
        )

        # -- Summary --
        total = t_cap + t_vlm + t_xport
        print_break()
        print_timing("total", total)

        if xlsx_path:
            shape = _xlsx_shape(xlsx_path)
            extras = []
            if "```" in raw_stripped:
                extras.append("fenced")
            hint = f"  ({shape})" + (f"  [{', '.join(extras)}]" if extras else "")
            print_ok(f"{xlsx_path}  {hint}")
        elif raw_stripped == "NO_TABLE":
            print_warn("no table detected in the selected region")
            print_tip("Make sure the area has a table with column headers")
        elif raw_stripped.startswith("ERROR:"):
            print_err(raw_stripped)
    except ValueError:
        print_warn("VLM returned unparseable output")
        print_tip("Try again, or check Ollama model status")
    except Exception as exc:
        print_err(f"unexpected error: {exc}")
        print_tip("If this persists, check the logs")


# Image extensions accepted by batch mode (mirrors tests/test_end_to_end.py)
_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".webp"}


def process_image_file(image_path: str, timestamp: str | None = None) -> str | None:
    """Analyze one image file and export it to XLSX.

    Prints a ``file`` stage (with the image name) followed by the shared
    ``vlm`` / ``export`` stages.  Errors are reported through the output
    helpers and never raised to the caller.

    Args:
        image_path: Path to a supported image file.
        timestamp:  Optional pre-generated timestamp (``YYYY-MM-DD_HHMMSS``).
                    When ``None`` a new one is generated per call.

    Returns:
        Absolute path to the generated XLSX file, or ``None`` when the
        image is missing / unsupported, the VLM reports an error or no
        table, or export fails.
    """
    try:
        path = Path(image_path)
        if not path.is_file():
            print_err(f"file not found: {image_path}")
            print_tip("Check the path and try again")
            return None
        if path.suffix.lower() not in _IMAGE_EXTS:
            print_err(f"unsupported image type: {image_path}")
            print_tip(f"Supported: {', '.join(sorted(_IMAGE_EXTS))}")
            return None

        ts = timestamp or datetime.now().strftime("%Y-%m-%d_%H%M%S")

        print_stage(f"file  {path.name}")
        image_bytes = path.read_bytes()
        xlsx_path, raw_stripped, t_vlm, t_xport = _analyze_and_export(
            image_bytes, ts
        )

        total = t_vlm + t_xport
        print_break()
        print_timing("total", total)

        if xlsx_path:
            shape = _xlsx_shape(xlsx_path)
            extras = []
            if "```" in raw_stripped:
                extras.append("fenced")
            hint = f"  ({shape})" + (f"  [{', '.join(extras)}]" if extras else "")
            print_ok(f"{xlsx_path}  {hint}")
            return xlsx_path
        if raw_stripped == "NO_TABLE":
            print_warn("no table detected in the image")
            print_tip("Make sure the image has a table with column headers")
        elif raw_stripped.startswith("ERROR:"):
            print_err(raw_stripped)
        return None
    except ValueError:
        print_warn("VLM returned unparseable output")
        print_tip("Try again, or check Ollama model status")
        return None
    except Exception as exc:
        print_err(f"unexpected error: {exc}")
        print_tip("If this persists, check the logs")
        return None


def _ensure_ollama(url: str, wait: int = 4) -> bool:
    """Check Ollama reachability; auto-start if missing.

    Returns True if API is reachable within *wait* seconds.
    """
    # Already reachable?
    try:
        req = urllib.request.Request(f"{url}/api/tags")
        urllib.request.urlopen(req, timeout=2)
        return True
    except Exception:
        pass

    # Try to start Ollama server in background (no console window)
    try:
        subprocess.Popen(
            ["ollama", "serve"],
            creationflags=subprocess.CREATE_NO_WINDOW,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            stdin=subprocess.DEVNULL,
        )
    except Exception:
        return False

    for _ in range(wait * 2):
        time.sleep(0.5)
        try:
            req = urllib.request.Request(f"{url}/api/tags")
            urllib.request.urlopen(req, timeout=1)
            return True
        except Exception:
            continue
    return False


def main_loop() -> None:
    """Print banner and enter the hotkey-polling loop."""
    # Check Ollama before showing the banner so the hotkey is ready immediately.
    with spinner("checking Ollama"):
        ok = _ensure_ollama(VLM_OLLAMA_URL)
    if ok:
        print_ok("ollama reachable")
    else:
        print_warn(f"ollama unreachable ({VLM_OLLAMA_URL})")
        print_tip("Auto-launch failed. Start manually:  ollama serve")
        console.print()

    print_banner(HOTKEY.title(), str(OUTPUT_DIR.resolve()), VLM_MODEL)

    try:
        while True:
            wait_for_hotkey(HOTKEY)
            time.sleep(DEBOUNCE)
            process_screenshot()
            print_rule()
    except KeyboardInterrupt:
        console.print("\n[green]exited[/green]")
