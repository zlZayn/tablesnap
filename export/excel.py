"""Excel export functionality."""

import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill

from core.config import OUTPUT_DIR

_HIGHLIGHT_FILL = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)


def export_to_excel(
    data: list[list[str]],
    output_dir: str | None = None,
    highlight_rows: list[int] | None = None,
) -> str:
    """Export a 2-D grid of text data to an Excel file.

    Creates a timestamped ``.xlsx`` file under *output_dir* (defaults to
    ``config.OUTPUT_DIR``).  Column widths are auto-fitted with a 50-char
    cap.

    Args:
        data:           2-D list where ``data[row][col]`` is the cell text.
        output_dir:     Override output directory.
        highlight_rows: 0-based row indices to highlight with yellow fill.

    Returns:
        Absolute path to the generated Excel file.

    Raises:
        OSError:  If the directory cannot be created or the file written.
        ValueError: If *data* is empty.
    """
    if not data:
        raise ValueError("data must not be empty")

    if output_dir is None:
        output_dir = str(OUTPUT_DIR)

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"{timestamp}.xlsx"
    filepath = output_path / filename

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Results"

    highlight = set(highlight_rows or [])

    for row_idx, row in enumerate(data, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            if (row_idx - 1) in highlight:
                cell.fill = _HIGHLIGHT_FILL

    # Auto-fit column widths (skip MergedCell entries)
    for column_cells in ws.columns:
        max_length = 0
        first = column_cells[0]
        if isinstance(first, MergedCell):
            continue
        col_letter = first.column_letter
        for cell in column_cells:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(str(filepath))
    return str(filepath)


def csv_to_excel(
    csv_text: str,
    output_dir: str | None = None,
) -> str:
    """Parse CSV text from VLM output and export to Excel.

    If the VLM's response is wrapped in a ```csv fence, the content
    inside the fence is used; otherwise the full text is treated as CSV.
    Empty lines are dropped.  Each line is parsed with :mod:`csv.reader`.

    Args:
        csv_text:    Raw text returned by the VLM.
        output_dir:  Override output directory (default config.OUTPUT_DIR).

    Returns:
        Absolute path to the generated Excel file.

    Raises:
        ValueError: If csv_text is empty after cleaning.
        OSError:    If file cannot be written.
    """
    # Try to extract CSV from a markdown code fence first
    match = re.search(r"```csv\s*\n(.+?)\n```", csv_text, re.DOTALL)
    raw = match.group(1) if match else csv_text.strip()

    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
    if not lines:
        raise ValueError("csv_text is empty after cleaning")

    parsed: list[list[str]] = []
    for line in lines:
        reader = csv.reader([line])
        parsed.append(next(reader))

    return export_to_excel(parsed, output_dir)
