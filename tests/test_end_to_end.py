"""End-to-end batch test: VLM analyze -> PSV parse -> XLSX export.

Iterates over all images in ``tests/test_table_pics/`` and runs the full
VLM pipeline on each one, saving results to ``tests/test_output/`` for
manual review.

Usage:
    uv run python tests/test_end_to_end.py             # run all (dump mode, auto-cleans old xlsx/json)
    uv run python tests/test_end_to_end.py --show      # show last results from _report.json (no VLM)
"""

import sys
import time
import json
from pathlib import Path

# Ensure project root AND tests dir are on sys.path before any local imports
_proj_root = Path(__file__).resolve().parent.parent
_tests_dir = Path(__file__).resolve().parent
sys.path.insert(0, str(_proj_root))
sys.path.insert(0, str(_tests_dir))

from tests.read_xlsx import xlsx_shape, print_xlsx_rows

# Force UTF-8 for terminal display (Chinese filenames / content)
try:
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
except Exception:
    pass

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = _proj_root

from core.config import TEST_IMAGES, TEST_OUTPUT


# ===================================================================
#  Display helpers
# ===================================================================

def _display_report(entries: list[dict]) -> None:
    """Pretty-print a saved test report (from _report.json)."""
    for entry in entries:
        print(f"=== {entry['image']} ===")
        status = entry.get("status", "?")
        shape = entry.get("xlsx_shape", "?")
        vlm_time = entry.get("vlm_time_s")
        psv_raw = entry.get("psv_raw")
        xlsx_content = entry.get("xlsx_content")

        if status == "ok":
            print(f"  VLM: {vlm_time}s  | XLSX: {shape}"
                  f"  | PSV len: {entry.get('psv_len', '?')}")

            # Print VLM raw output
            if psv_raw:
                print("  --- VLM raw output ---")
                for line in psv_raw.splitlines():
                    print(f"  |{line}")
                print("  ---")

            # Print XLSX content from JSON snapshot
            if xlsx_content:
                print(f"  --- XLSX ({entry.get('xlsx_path', '?')}) ---")
                for row in xlsx_content[:30]:
                    print(f"    {row}")
                if len(xlsx_content) > 30:
                    print(f"    ... ({len(xlsx_content) - 30} rows hidden)")
        else:
            print(f"  STATUS: {status}", end="")
            if entry.get("error"):
                print(f"  | ERROR: {entry['error']}", end="")
            print()
        print()


def _print_header() -> None:
    hdr = f"{'Image':40s} {'Size':>8s} {'VLM(ms)':>8s} {'Rows':>6s} {'Status':12s}"
    print(hdr)
    print("-" * 80)


def _print_row(r: dict, rows_str: str) -> None:
    vlm_ms = f"{r['vlm_time_s']*1000:.0f}" if r["vlm_time_s"] else "-"
    name = r["image"]
    # Truncate long names for table alignment
    if len(name) > 38:
        name = name[:35] + "..."
    print(f"{name:40s} {r['size_bytes']:>8d} {vlm_ms:>8s} {rows_str:>6s} {r['status']:12s}")
    if r["status"] in ("error", "parse_error", "export_error"):
        print(f"  {'':40s} error: {r['error']}")


# ===================================================================
#  Core test logic
# ===================================================================

def run_one(image_path: Path) -> dict:
    """Run the full VLM->PSV->XLSX pipeline on a single image.

    Returns a dict with stats and the output XLSX path (or error).
    """
    result = {
        "image": image_path.name,
        "size_bytes": image_path.stat().st_size,
        "status": "ok",
        "vlm_time_s": None,
        "psv_len": None,
        "xlsx_path": None,
        "error": None,
    }

    from vlm.client import OllamaClient
    from export.xlsx import psv_to_xlsx
    from core.config import VLM_OLLAMA_URL, VLM_MODEL, VLM_TIMEOUT

    client = OllamaClient(
        base_url=VLM_OLLAMA_URL,
        model=VLM_MODEL,
        timeout=VLM_TIMEOUT,
    )

    image_bytes = image_path.read_bytes()
    t0 = time.perf_counter()
    psv_text = client.analyze(image_bytes)
    t1 = time.perf_counter()
    result["vlm_time_s"] = round(t1 - t0, 2)

    if psv_text.startswith("ERROR:"):
        result["status"] = "error"
        result["error"] = psv_text
        return result

    if psv_text.strip() == "NO_TABLE":
        result["status"] = "no_table"
        result["psv_len"] = 0
        return result

    result["psv_raw"] = psv_text  # full VLM output for dump display
    result["psv_len"] = len(psv_text)

    try:
        xlsx_path = psv_to_xlsx(psv_text, output_dir=str(TEST_OUTPUT))
        result["xlsx_path"] = xlsx_path
    except ValueError as e:
        result["status"] = "parse_error"
        result["error"] = str(e)
    except Exception as e:
        result["status"] = "export_error"
        result["error"] = f"{type(e).__name__}: {e}"

    return result


def _clean_output() -> None:
    """Remove old XLSX and JSON files from test output."""
    count = 0
    for f in list(TEST_OUTPUT.glob("*.xlsx")) + list(TEST_OUTPUT.glob("*.json")):
        f.unlink()
        count += 1
    if count:
        print(f"(cleaned {count} stale files)\n")


def run_all() -> list[dict]:
    """Run the pipeline on every image in TEST_IMAGES."""
    TEST_OUTPUT.mkdir(parents=True, exist_ok=True)
    _clean_output()

    images = sorted(TEST_IMAGES.glob("*.*"))
    valid_exts = {".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".webp"}
    images = [i for i in images if i.suffix.lower() in valid_exts]

    if not images:
        print(f"No test images found in {TEST_IMAGES}")
        print("Place screenshot images there and re-run.")
        return []

    _print_header()
    results = []

    for img in images:
        r = run_one(img)
        rows_str = xlsx_shape(r["xlsx_path"]) if r["xlsx_path"] else "-"
        _print_row(r, rows_str)
        results.append(r)

    ok_count = sum(1 for r in results if r["status"] == "ok")
    fail_count = sum(1 for r in results if r["status"] != "ok")
    print("-" * 80)
    print(f"Total: {len(results)}  OK: {ok_count}  Failed: {fail_count}")
    print(f"\nOutput: {TEST_OUTPUT.resolve()}")

    # Persist report
    report_path = TEST_OUTPUT / "_report.json"
    report_data = []
    for r in results:
        entry = {
            "image": r["image"],
            "size_bytes": r["size_bytes"],
            "status": r["status"],
            "vlm_time_s": r["vlm_time_s"],
            "psv_len": r["psv_len"],
            "xlsx_path": r["xlsx_path"],
            "error": r["error"],
        }
        # Include VLM raw output for dump replay
        if r.get("psv_raw"):
            entry["psv_raw"] = r["psv_raw"]
        # Snapshot XLSX content so --show works without re-loading file
        if r["xlsx_path"] and r["status"] == "ok":
            try:
                from openpyxl import load_workbook
                wb = load_workbook(r["xlsx_path"])
                ws = wb.active
                if ws is not None:
                    entry["xlsx_content"] = [
                        [str(c) if c is not None else "" for c in row]
                        for row in ws.iter_rows(values_only=True)
                    ]
                    entry["xlsx_shape"] = f"{ws.max_row}r x {ws.max_column}c"
            except Exception:
                pass
        report_data.append(entry)
    report_path.write_text(
        json.dumps(report_data, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"Report: {report_path.resolve()}")

    return results


def show_last() -> None:
    """Display results from the last saved report (no VLM calls)."""
    report_path = TEST_OUTPUT / "_report.json"
    if not report_path.exists():
        print("No saved report found. Run without --show first.")
        return
    entries = json.loads(report_path.read_text("utf-8"))
    _display_report(entries)


# ===================================================================
#  CLI
# ===================================================================

def main():
    args = [a.lower() for a in sys.argv[1:]]

    if "--show" in args:
        show_last()
        return

    # Default: run all with dump
    results = run_all()
    print()
    for r in results:
        print(f"=== {r['image']} ===")
        print(f"  status: {r['status']}")
        if r["vlm_time_s"]:
            print(f"  vlm:    {r['vlm_time_s']}s")
        print()

        # Print raw VLM output
        raw = r.get("psv_raw", "")
        if raw:
            print("  --- VLM raw output ---")
            for line in raw.splitlines():
                print(f"  |{line}")
            print("  ---")
        print()

        # Print XLSX content
        if r["status"] == "ok" and r["xlsx_path"]:
            xp = r["xlsx_path"]
            print(f"  --- XLSX ({xp}) ---")
            from openpyxl import load_workbook
            wb = load_workbook(xp)
            ws = wb.active
            if ws is not None:
                print(f"  sheet: {ws.title}  ({ws.max_row}r x {ws.max_column}c)")
                for row in ws.iter_rows(
                    min_row=1,
                    max_row=min(ws.max_row or 0, 30),
                    values_only=True,
                ):
                    print(f"    {list(row)}")
            print()
        elif r.get("error"):
            print(f"  error: {r['error']}")
            print()


if __name__ == "__main__":
    main()
