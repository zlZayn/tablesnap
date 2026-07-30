"""Read an XLSX file and print as text table.

Usage:
    uv run python tests/read_xlsx.py path/to/file.xlsx
    uv run python tests/read_xlsx.py path/to/file.xlsx --rows 5
"""

import sys
from pathlib import Path

from openpyxl import load_workbook


def xlsx_shape(xlsx_path: str) -> str:
    """Return 'rowsxcols' string for a given XLSX file."""
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active
        if ws is not None:
            return f"{ws.max_row}x{ws.max_column}"
        return "?"
    except Exception:
        return "?"


def print_xlsx_rows(ws, max_rows: int = 30) -> None:
    """Print worksheet rows as lists (reusable, no header)."""
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row or 0, max_rows),
        values_only=True,
    ):
        values = []
        for v in row:
            if v is None:
                values.append("")
            else:
                s = str(v)
                if len(s) > 60:
                    s = s[:57] + "..."
                values.append(s)
        print(f"  {values}")


def dump_xlsx(xlsx_path: str, max_rows: int | None = None) -> None:
    """Print an openpyxl workbook as rows of lists."""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    if ws is None:
        print(f"No active sheet in {xlsx_path}")
        return

    nrows = ws.max_row or 0
    ncols = ws.max_column or 0
    limit = nrows if max_rows is None else min(nrows, max_rows)

    print(f"File : {xlsx_path}")
    print(f"Sheet: {ws.title}  ({nrows}r x {ncols}c)")
    print()

    print_xlsx_rows(ws, max_rows=limit)

    if max_rows is not None and nrows > max_rows:
        print(f"  ... ({nrows - max_rows} rows hidden)")


def main() -> None:
    args = [a.lower() for a in sys.argv[1:]]

    # Extract positional paths (everything before --rows / -n)
    paths: list[str] = []
    max_rows: int | None = None
    skip = False
    for i, a in enumerate(sys.argv[1:], start=1):
        if skip:
            skip = False
            continue
        if a in ("--rows", "-n") and i < len(sys.argv):
            try:
                max_rows = int(sys.argv[i + 1])
            except ValueError:
                print(f"Invalid row count: {sys.argv[i+1]}")
                sys.exit(1)
            skip = True
        elif not a.startswith("-"):
            paths.append(a)

    if not paths:
        print((__doc__ or "").strip())
        sys.exit(1)

    for p in paths:
        if not Path(p).exists():
            print(f"File not found: {p}")
            continue
        dump_xlsx(p, max_rows=max_rows)
        print()


if __name__ == "__main__":
    main()
